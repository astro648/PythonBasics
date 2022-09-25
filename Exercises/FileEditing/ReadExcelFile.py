# need pip modules xlrd and openpyxl
import xlrd
import openpyxl

# define variable to load the dataframe
dataframe = openpyxl.load_workbook('1MonthOfLaunches.xlsx')

# define variable to read sheet
dataframe1 = dataframe.active

# read cell values
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)
